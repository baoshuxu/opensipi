import re
import io
import os
import time
import pytz
import json
import pickle
import base64
import openpyxl
import mimetypes
import pandas as pd
from datetime import datetime, timedelta
from io import StringIO
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

import logging

class GExecuteError(Exception):
    pass

def suppress_google_api_logging():
  """Suppresses logging from various Google API client libraries."""
  # Google Sheets and general API client
  logging.getLogger("googleapiclient").setLevel(logging.ERROR)

  # Google Authentication
  logging.getLogger("google.auth").setLevel(logging.ERROR)
  logging.getLogger("requests").setLevel(logging.ERROR)
  logging.getLogger("urllib3").setLevel(logging.ERROR)
  logging.getLogger("urllib3.connectionpool").setLevel(logging.ERROR)

  # Google Docs
  logging.getLogger("googleapiclient.discovery").setLevel(logging.ERROR)
  logging.getLogger("googleapiclient.http").setLevel(logging.ERROR)

  # Google Drive
  logging.getLogger("googleapiclient.discovery").setLevel(logging.ERROR)
  logging.getLogger("googleapiclient.http").setLevel(logging.ERROR)

  # Gmail
  logging.getLogger("googleapiclient.discovery").setLevel(logging.ERROR)
  logging.getLogger("googleapiclient.http").setLevel(logging.ERROR)

def col_key(cell):
# Example usage:
# cells = ['A11', 'A9', 'AA1', 'Z5', 'B2', 'AB10', 'Z1', 'AA10']
# sorted_cells = sorted(cells, key=col_key)
  match = re.match(r'^([A-Z]+)(\d+)$', cell.upper())
  if not match:
      return (float('inf'), float('inf'))  # invalid cells last
  col_str, row = match.groups()
  col_num = 0
  for char in col_str:
      col_num = col_num * 26 + (ord(char) - ord('A') + 1)
  return (col_num, int(row))


# Call the function to suppress logs
suppress_google_api_logging()

API_REQUESTS = {
    "count": 60,
    "time": time.time() + 60,
}  # gsheet quota 60 operations per 60 seconds
API_SERVICES = {
    "sheets": {
        "version": "v4",
        "discovery": "https://sheets.googleapis.com/$discovery/rest?version=v4",
        "scopes": "https://www.googleapis.com/auth/spreadsheets",
    },
    "docs": {
        "version": "v1",
        "discovery": "https://docs.googleapis.com/$discovery/rest?version=v1",
        "scopes": "https://www.googleapis.com/auth/documents.readonly",
    },
    "gmail": {
        "version": "v1",
        "discovery": "https://gmail.googleapis.com/$discovery/rest?version=v1",
        "scopes": "https://mail.google.com/",
    },
    "drive": {
        "version": "v3",
        "discovery": "https://www.googleapis.com/discovery/v1/apis/drive/v3/rest",
        "scopes": "https://www.googleapis.com/auth/drive",
    },
}


def gservice(api_name):
  global API_SERVICES
  credentials = "support/credentials.json"
  if api_name in API_SERVICES and API_SERVICES[api_name].get("build", None):
    return API_SERVICES[api_name]["build"]

  # print(credentials, api_name, api_version, scopes, sep='-')
  cred = None
  pickle_file = f"support/credentials.pickle"
  if os.path.exists(pickle_file):
    with open(pickle_file, "rb") as token:
      cred = pickle.load(token)

  if not cred or not cred.valid:
    if cred and cred.expired and cred.refresh_token:
      cred.refresh(Request())
    else:
      API_SCOPES = [d["scopes"] for (_, d) in API_SERVICES.items()]
      flow = InstalledAppFlow.from_client_secrets_file(credentials, API_SCOPES)
      # flow = InstalledAppFlow.from_client_secrets_file(credentials, scopes)
      cred = flow.run_local_server()

    with open(pickle_file, "wb") as token:
      pickle.dump(cred, token)

  try:
    api_verssion = API_SERVICES[api_name]["version"]
    service_url = API_SERVICES[api_name]["discovery"]
    service = build(api_name, api_verssion, credentials=cred, discoveryServiceUrl=service_url)
    print(api_name, "service created successfully")
    API_SERVICES[api_name].update({"build": service})
    return service
  except Exception as e:
    print(f"Unable to connect. {e}")
    return None


def gexecute(req, retries=5, delay=5):
  # check quota
  global API_REQUESTS
  if API_REQUESTS["count"] > 0:
    API_REQUESTS["count"] -= 1
    if req.method in ["PUT", "POST"]:
      time.sleep(1)  # enforce 1 wirte/sec, otherwise quota error if API_REQUESTS['count'] > 0 :
  else:
    # API_REQUESTS['count'] = 60
    # if 'GET' == req.method:
    while time.time() < API_REQUESTS["time"]:
      time.sleep(1)
    API_REQUESTS.update({"count": 60, "time": time.time() + 60})

  # retries when http error
  for attempt in range(retries):
    try:
      return req.execute()
    except TimeoutError as err:
      if attempt < retries - 1:
        print(f"gexecute: Timeout at attempt {attempt}. Retrying in {delay} seconds...")
        time.sleep(delay)
        delay *= 2
      else:
        raise
    except HttpError as err:
      if err.resp.status == 502:
        print(f"gexecute: Bad Gateway error (502) encountered. Retrying in {delay} seconds...")
        time.sleep(delay)
        delay *= 2
      elif err.resp.status == 429:  # Rate limit error
        print(f"gexecute: Rate limit error(429), retrying in {delay} seconds...")
        time.sleep(delay)
        delay *= 2
      elif err.resp.status == 500:
        print(f"gexecute: Internal error (500) encountered. Retrying in {delay} seconds...")
        time.sleep(delay)
        delay *= 2
      else:
        raise
  print(f"Max retries ({retries}) exceeded for http_request {req.method}.")
  raise GExecuteError("gexecute failure")

class gsheet:

  def __init__(self, url, xls=None):
    self.service = gservice("sheets")
    self.book_id = re.search(".+/spreadsheets/d/(.+)/edit.*", url).group(1)
    self.page_id = {}  # pages and its sheetId
    self.page_mt = {}  # pages and its modify time
    self.__xlsx = xls  # local cache
    self.requests = []  # flush buffer
    self._flush = True
    self.pages()  # always read pages first

  def flush(self, yes=None):
    if yes is None:
      return self._flush
    self._flush = yes
    if self._flush:
      gexecute(self.service.spreadsheets().batchUpdate(spreadsheetId=self.book_id, body={"requests": self.requests}))
    else:
      self.requests = []

  def _to_columns(self, content={}):
    """content translation to columns (list of list) by moving key to data as df.to_dict() may drop duplicated columns"""
    if isinstance(content, pd.DataFrame):
      df = pd.DataFrame({i: [k] for (i, k) in enumerate(content.columns)})
      cnt = content.copy()  # do not modify the input argmument content!
      cnt.columns = df.columns
      cnt.fillna("", inplace=True)
      cnt = cnt.infer_objects(copy=False)
      df = pd.concat([df, cnt], ignore_index=True)
      data = df.to_dict(orient="list")
    elif isinstance(content, dict):
      data = {i: [k] + v for i, (k, v) in enumerate(content.items())}
    else:
      data = {}
    # prepare data: dictionay to list of columns
    columns = [[]] * len(data)
    for i, v in data.items():
      columns[i] = v
      for j, ele in enumerate(columns[i]):
        if isinstance(ele, list):
          columns[i][j] = ",".join(ele)
    return columns

  def isrange(self, string):
    rng = re.split("[!:]", string)
    return rng and (rng[0] in self.pages())

  def pages(self, tab=None, add=None, delete=None):
    """list each page and its Id"""
    refresh = True if (add or delete or len(self.page_id) == 0) else False  # report pages if refresh
    if bool(delete) and (delete in self.page_id):
      # delete named ranges first
      req = []
      r = gexecute(self.service.spreadsheets().get(spreadsheetId=self.book_id))
      if r is None:
        return
      sheet_names = [e["properties"]["title"] for e in r.get("sheets")]

      named_rs = r.get("namedRanges", [])
      named = []
      for rs in named_rs:
        if self.page_id[delete] == rs["range"]["sheetId"]:
          named.append(rs["namedRangeId"])
      if len(named):
        req += [{"deleteNamedRange": {"namedRangeId": i}} for i in named]
      req += [{"deleteSheet": {"sheetId": self.page_id[delete]}}]
      if delete in sheet_names:  # someone may delete the sheet after self.page_id update
        gexecute(self.service.spreadsheets().batchUpdate(spreadsheetId=self.book_id, body={"requests": req}))
      self.page_id.pop(delete)
    if bool(add) and not (add in self.page_id):
      body = {"requests": [{"addSheet": {"properties": {"title": add}}}]}
      gexecute(self.service.spreadsheets().batchUpdate(spreadsheetId=self.book_id, body=body))
    if refresh:
      r = gexecute(self.service.spreadsheets().get(spreadsheetId=self.book_id))
      for item in r.get("sheets"):
        sheetName = item["properties"]["title"]  # .strip()
        self.page_id.update({sheetName: item["properties"]["sheetId"]})
        self.page_mt.update({sheetName: None})  # todo
        # self.page_mt.update({sheetName: item['properties']['updatedTime']})
    if tab is None:
      return self.page_id
    elif tab in self.page_id:
      return self.page_id[tab]
    return None

  def copy(self, src=None, tgt=None):
    """copy a src page to tgt page in the same gsheet"""
    if src and src in self.pages():
      self.pages(delete=tgt)
      body = {"destinationSpreadsheetId": self.book_id}
      res = gexecute(
          self.service.spreadsheets().sheets().copyTo(spreadsheetId=self.book_id, sheetId=self.page_id[src], body=body)
      )
      if "sheetId" in res:
        tgt_id = res["sheetId"]
        req = [
            {
                "updateSheetProperties": {
                    "properties": {"sheetId": tgt_id, "title": tgt},
                    "fields": "title",
                }
            }
        ]
        ans = gexecute(self.service.spreadsheets().batchUpdate(spreadsheetId=self.book_id, body={"requests": req}))
        if "replies" in ans and len(ans["replies"]) > 0:
          self.page_id.update({tgt: tgt_id})

  def clear(self, tab, options={}):
    """clear a page"""
    req = []
    sheetId = self.pages(tab)
    if "named_ranges" in options:
      res = gexecute(self.service.spreadsheets().get(spreadsheetId=self.book_id))
      # res= gexecute(self.service.spreadsheets().get(spreadsheetId=self.book_id, ranges=[], includeGridData=False))
      named_ranges = res.get("namedRanges", [])
      invalid_names = [nr["name"] for nr in named_ranges if nr["range"].get("getA1Notation") == "#REF!"]
      req += [{"deleteNamedRange": {"namedRangeId": name}} for name in invalid_names]
    if sheetId is not None:
      req += [
          {
              "updateCells": {
                  "range": {"sheetId": sheetId},
                  "fields": "userEnteredValue",
              }
          }
      ]
      if "named_ranges" in options:
        sheet_names = [nr["name"] for nr in named_ranges if nr["range"]["sheetId"] == sheetId]
        req += [{"deleteNamedRange": {"namedRangeId": name}} for name in sheet_names]
    else:
      self.pages(add=tab)
    if req:
      if self.flush():
        gexecute(self.service.spreadsheets().batchUpdate(spreadsheetId=self.book_id, body={"requests": req}))
      else:
        self.requests.extend(req)

  def remove(self, tabs=[], options={}):
    """remove pages :
    remove(tabs=['page1','page2'], options={'hidden': True})
    """
    if len(tabs) < 1 and len(options) < 1:
      return
    rm_sheetIds = [self.page_id[tab] for tab in tabs]
    if "hidden" in options:
      res = gexecute(self.service.spreadsheets().get(spreadsheetId=self.book_id))
      sheets = res.get("sheets", [])
      hidden_sheet_ids = []
      for sheet in sheets:
        sheet_properties = sheet.get("properties", {})
        if sheet_properties.get("hidden", False):
          hidden_sheet_ids.append(sheet_properties["sheetId"])
      rm_sheetIds += hidden_sheet_ids
    if rm_sheetIds:
      req = [{"deleteSheet": {"sheetId": sheet_id}} for sheet_id in rm_sheetIds]
      gexecute(self.service.spreadsheets().batchUpdate(spreadsheetId=self.book_id, body={"requests": req}))

  def columns(self, tab=""):
    """read tab with headers to dictionary of columns"""
    if tab and tab in self.pages():
      r = gexecute(self.service.spreadsheets().values().get(spreadsheetId=self.book_id, majorDimension="COLUMNS", range=tab))
      return r.get("values", {})
    else:
      print(f"gsheet.columns(): page {tab} not found")
      return []

  def rows(self, tab="", visible=None):
    """read tab with headers to dictionary of rows"""
    if tab and tab in self.pages():
      r = gexecute( self.service.spreadsheets().values().get(spreadsheetId=self.book_id, majorDimension="ROWS", range=tab) )
    else:
      print(f"ghsheet.rows(): page {tab} not found")
      return []
    if visible:
      shown = self.visible(tab, 0)["shown"]
      return [e for (i, e) in enumerate(r["values"]) if i in shown]
    else:
      return r["values"]

  def cells(self, range="Sheet1!A1:Z26"):
    """read cells. native google sheet api return is list of rows"""
    if self.isrange(range) is False:
      print(f"gsheet.cells(): range {range} is not found")
      return []
    z = re.search(r"([^!]+)!(\w+):?(\w*)", range)
    if z:
      r = gexecute(self.service.spreadsheets().values().get(spreadsheetId=self.book_id, majorDimension="ROWS", range=range) )
      return r["values"]
    else:
      return []

  def text(self, tab="", visible=None):
    """read tab to comma seperated text"""
    rows = self.rows(tab, visible)
    txt = [""] * len(rows)
    ncols = [len(x) for x in rows]
    for i in range(len(rows)):
      if ncols[i] < max(ncols):
        rows[i] += [""] * (max(ncols) - ncols[i])
      txt[i] = ",".join(rows[i])
    return txt

  def df(self, tab="", header_row=0, visible=None, cached=False):
    """read a page into a datafram, assiming 1st header as df columns"""
    if not tab:
      return pd.DataFrame()
    df = self.retrieve(tab, header_row) if cached else None
    if df is None:  # try gsheet read if retrieve failed
      d = self.read(tab, header_row)
      if not bool(d):
        return pd.DataFrame()
      df = pd.DataFrame.from_dict(d, dtype="str")
      df.fillna("", inplace=True)
    if visible is None:
      return df if df is not None else pd.DataFrame()
    if visible:
      x = self.visible(tab)["hidden"]
      df.drop([i for i in df.index if header_row + 1 + i in x], inplace=True)
      df.reset_index(drop=True, inplace=True)
      return df if df is not None else pd.DataFrame()

  def table(self, tab="", skip_blank_col=0, data=None):
    # talbes are always seperated by blank rows
    # name in firt col of non-empty row, header in next row and content in rest rows
    # rows in table droped if cell is blank at skip_blank_col
    # visibility attached as a column
    if not tab: 
      print('no page name specified for gsheet.table()')
      return
    
    if data: # write multiple dfs
      self.pages(add=tab)
      cells = []
      for (name, df) in data.items():
        cells.append([name])
        cells.extend( list(map(list, zip(*self._to_columns(df)))) )
        cells.append([''])
      max_cols = max([len(r) for r in cells])
      params = {
          'spreadsheetId': self.book_id,
          'valueInputOption': "USER_ENTERED",
          'range': f"{tab}!A1:{chr(max_cols+64)}{len(cells)}",
          'body': {"majorDimension": "ROWS", "values": cells}
      }
      gexecute(self.service.spreadsheets().values().update(**params) )
      return
    
    if tab not in self.pages():
      print('No page name specified to read tables from')
      return
    
    visbility = self.visible(tab)
    df_all = self.df(tab, header_row=None)
    tbl_dict = {}

    data_tbl, tbl_title, headers = [], None, None

    for idx, row in df_all.iterrows():
      # Consider row blank if all cells are empty
      if all(c == "" for c in row):
        # Blank row, finalize the current table
        if data_tbl:
          df = pd.DataFrame(data_tbl)
          df.columns = headers
          tbl_dict[tbl_title] = df
        elif headers:
          tbl_dict[tbl_title] = pd.DataFrame(columns=headers)
        elif tbl_title:
          tbl_dict[tbl_title] = pd.DataFrame()
        data_tbl, tbl_title, headers = [], None, None
      elif tbl_title is None:
        tbl_title = row.iloc[0] if row.iloc[0] != "" else None
      elif headers is None:
        headers = [c for c in row if c != ""] if any(c != "" for c in row) else None
      else:
        trimmed_row = row[: len(headers)]  # Trim to match header length
        data_tbl.append(trimmed_row)

    # Finalize any remaining table
    if data_tbl and tbl_title:
      df = pd.DataFrame(data_tbl)
      df.columns = headers
      tbl_dict[tbl_title] = df

    # append 'hide' column and drop rows if cell at skip_blank_col is empty
    for k, df in tbl_dict.items():
      df["hide"] = [False if i in visbility["shown"] else True for i in df.index]
      if isinstance(skip_blank_col, int) and skip_blank_col >= 0 and skip_blank_col < len(headers):
        df = df[df.iloc[:, skip_blank_col] != ""]  # rows droped if first element empty
      tbl_dict[k] = df

    return tbl_dict

  def c2df(self, range="Sheet1!A1:Z999", header="infer"):
    """ data in cells to dataframe """
    rows = self.cells(range)
    txt = [""] * len(rows)
    ncols = [len(x) for x in rows]
    for i, row in enumerate(rows):
      if ncols[i] < max(ncols):
        row += [""] * (max(ncols) - ncols[i])
      txt[i] = ",".join(rows[i])
    if bool(txt):
      csvstring = StringIO("\n".join(txt))
      data = pd.read_csv(
          csvstring,
          dtype="str",
          sep=r"\s*,\s*",
          header=header,
          keep_default_na=False,
          engine="python",
      )
      return data
    else:
      return None

  def read(self, tab="", k=0, blurry=False):
    """ read tab with headers to dictionary of columns """
    content, columns = ({}, self.columns(tab))
    nrows = [len(x) for x in columns]
    for i in range(len(columns)):
      if nrows[i] < max(nrows):
        columns[i] += [""] * (max(nrows) - nrows[i])
      columns[i] = [s.strip() for s in columns[i]]
      if k is None:
        content[chr(ord("A") + i)] = columns[i]
      elif isinstance(k, int):
        content[columns[i][k]] = columns[i][k + 1 :]
    if blurry:
      content["visible"] = self.visible(tab, k)["visible"]
    return content

  def scan(self, tab="", filter="simulate=TRUE,visible=yes", cached=False):
    """ read tab with headers to dictionary of columns, filtered """
    if not (tab and tab in self.pages()):
      return {}
    df = self.retrieve(tab) if cached else None
    if df is not None:
      content = df.to_dict(orient="list")
    else:  # try gsheet incase retrieve failure
      content = self.read(tab)
    k0 = next(iter(content))
    operator = ["==", ">=", "<=", ">", "<", "="]
    filters = filter.replace(" ", "").split(",")
    for filt in filters:
      compare = [x in filt for x in operator]
      if len(compare) < 1:
        continue
      op = operator[compare.index(True)]  # only first True, i.e, '>=' picked rather than '>'
      try:
        var, yes = re.sub(op, " ", filt, count=0).split()
      except:
        continue
      if var == "simulate":
        if re.match(r"^ALL$", yes, re.IGNORECASE):
          sel = range(len(content[k0]))
        else:
          sel = [i for (i, x) in enumerate(content[var]) if x == yes]
        for k, col in content.items():
          content[k] = [e for (i, e) in enumerate(col) if i in sel]
      elif var == "visible":
        if re.match(r"ALL", yes, re.IGNORECASE):
          shown = range(len(content[k0]))
        else:
          shown = self.visible(tab, 1)["shown"]
        for k, col in content.items():
          content[k] = [e for (i, e) in enumerate(col) if i in shown]
      elif var in content:
        satisfied = [i for (i, x) in enumerate(content[var]) if eval(f"{x}{op}{yes}")]
        for k, col in content.items():
          content[k] = [e for (i, e) in enumerate(col) if i in satisfied]

    return content

  def write(self, tab, content, rows=(), cols=()):
    """write data to tab range from A1"""
    if not tab in self.pages():
      self.pages(add=tab)
    self.clear(tab=tab)
    rng = f"{tab}!A1"
    if bool(rows) and len(rows) == 2:
      m = "A" + f"{(max(1,rows[0]))}"
      n = chr(ord("A") + len(columns) - 1) + f"{rows[1]}"
      rng = f"{tab}!{m}:{n}"
    if bool(cols) and len(cols) == 2:
      m = chr(ord("A") + (max(1, cols[0]))) + "1"
      n = chr(ord("A") + (max(2, cols[1]))) + f"{(max(1,columns[0]))}"
      rng = f"{tab}!{m}:{n}"
    content_arr = content if isinstance(content, list) else [content]
    M,N,H = 0,0,[]
    for content in content_arr:
      columns = self._to_columns(content)
      if len(content_arr)>1:
        m = "A" + f"{M +1}"
        n = chr(ord("A") + len(columns) - 1) + f"{len(columns[0])+M}"
        rng = f"{tab}!{m}:{n}"
    # write to tab
      values = {"majorDimension": "COLUMNS", "values": columns}
      params = {
          'spreadsheetId': self.book_id,
          'valueInputOption': "USER_ENTERED",
          'range': rng,
          'body': values
      }
      gexecute(self.service.spreadsheets().values().update(**params) )
      H += [ M, M+1]
      M += len(columns[0]) +1
      N = max(N, len(columns))
    shape = (5 * (1 + M // 5), 3 * (1 + N // 3))
#    self.render(tab, {"header": H, "align": None, "clip": shape})
    self.render(tab, {"header": None, "align": None, "clip": shape})

  def cache(self, tab, rows):
    """write data to local file (self.__xlsx), data in rows format"""
    if self.__xlsx is None:
      return
    try:
      workbook = openpyxl.load_workbook(self.__xlsx)
    except FileNotFoundError:
      workbook = openpyxl.Workbook()
    except Exception as e:
      print(f"gsheet.cache() error: {e}")
      return
    # always a overwrite by deleting old tabs if already there
    try:
      if tab in workbook.sheetnames:
        sheet = workbook[tab]
        sheet.delete_rows(1, sheet.max_row)
      else:
        sheet = workbook.create_sheet(title=tab)
      for row in rows:
        sheet.append(row)
      workbook.save(self.__xlsx)
    except PermissionError:
      print(f"?? {self.__xlsx} permission denied. {tab} not updated.")
      return
    except Exception as e:
      print(f"gsheet.cache() error: {e}")

  def retrieve(self, tab, header_row=0):
    """write data to local file (self.__xlsx), data in rows format"""
    if self.__xlsx is None:
      return None
    try:
      df = pd.read_excel(self.__xlsx, sheet_name=tab, header=header_row, engine="openpyxl")
      return df
    except PermissionError:
      print(f"?? {self.__xlsx} permssion denied. no data read from {tab}")
      return None
    except Exception as e:
      print(f"gsheet.retrieve() error: {e}")

  def update(self, tab, content, **kwargs):
    """clean and refresh tab with data, plus sort of columns"""
    styles = "merge border align header clip hide zebra".split()
    self.pages(delete=tab, add=tab)
    columns = self._to_columns(content)
    rows = list(map(list, zip(*columns)))

    # from operator import itemgetter
    # sort_cols = kwargs.get("sort")
    # if sort_cols and all(x in rows[0] for x in sort_cols):
    #   indices = [rows[0].index(i) for i in sort_cols]
    #   rows[1:] = sorted(rows[1:], key=itemgetter(*indices))

    if "sort" in kwargs:
      sort_cols = kwargs["sort"]
      if bool(sort_cols) and all([x in rows[0] for x in sort_cols]):
        k = ",".join([f"x[{rows[0].index(i)}]" for i in sort_cols])
        rows[1:] = eval(f"sorted(rows[1:],key=lambda x: ({k}), reverse=False)")
    options = {k: v for (k, v) in kwargs.items() if k in styles}

    # cache udpates to local xls, fear failing of gsheet write
    self.cache(tab, rows)

    # write to page
    cells = f"{tab}!A1"
    values = {"majorDimension": "ROWS", "values": rows}
    shape = (5 * (1 + len(rows) // 5), 3 * (1 + len(columns) // 3))
    params = {
      'spreadsheetId': self.book_id,
      'valueInputOption': "USER_ENTERED",
      'range': cells,
      'body': values
    }   
    gexecute(self.service.spreadsheets().values().update(**params) )
    self.flush(False)
    # self.render(tab, {"header": [0, 1], "align": None, "clip": shape, "hide": hide})
    options["header"] = options.get("header", [0, 1])
    options["hide"] = options.get("hide", False)
    options["clip"] = options.get("clip", shape)
    options["align"] = options.get("align", None)
    self.render(tab, options)
    self.flush(True)

  def set(self, tab, cell="A1", text=""):
    """write just one cell or a range start at cell"""
    loc = re.search(r"([A-Z]+)(\d+)", cell).groups()

    if not tab in self.pages():
      self.pages(add=tab)
    if isinstance(text, list):
      rows, cols = len(text[0]), len(text)
      tgt = chr(ord(loc[0]) + cols - 1) + str(int(loc[1]) + rows - 1)
      cells = f"{tab}!{cell}:{tgt}"
      columns = text
    elif isinstance(text, str):
      cells = f"{tab}!{cell}"
      columns = [[text]]
    values = {"majorDimension": "COLUMNS", "values": columns}
    params = {
      'spreadsheetId': self.book_id,
      'valueInputOption': "USER_ENTERED",
      'range': cells,
      'body': values
    }   
    gexecute(self.service.spreadsheets().values().update(**params) )

  def visible(self, tab="", start_row=0):
    """find rows visible or hidden"""
    fields = "sheets(data(rowMetadata(hiddenByFilter,hiddenByUser)))"
    res = gexecute(self.service.spreadsheets().get(spreadsheetId=self.book_id, ranges=tab, fields=fields))
    metadata = res["sheets"][0]["data"][0]["rowMetadata"]
    filtered = {"shown": [], "hidden": [], "visible": []}
    for i, r in enumerate(metadata[start_row:]):
      hide = r.get("hiddenByFilter", False) or r.get("hiddenByUser", False)
      filtered["hidden" if hide else "shown"].append(i)
      filtered["visible"].append(True if i in filtered["hidden"] else False)
    return filtered

  def render(self, tab, how):
    """render with merge/border/align/header/clip/zebra/dropdn/hide"""
    if tab not in self.pages():
      return
    if isinstance(how, tuple):
      how = self.style(tab, how)
      return
    for keys, param in how.items():
      if "merge" in keys:
        self.merge(tab, param)
      elif "border" in keys:
        self.border(tab, param)
      elif "align" in keys:
        self.align(tab, param)
      elif "header" in keys:
        self.header(tab, param)
      elif "clip" in keys:
        self.clip(tab, param)
      elif "zebra" in keys:
        self.zebra(tab, param)
      elif "dropdn" in keys:
        self.dropdn(tab, param)
      elif "hide" in keys:
        self.hide(tab, param)

  def hide(self, tab="", yes=True):
    """hide a page in gsheet"""
    sheetId = self.pages(tab)
    if sheetId is None:
      return
    req = [
        {
            "updateSheetProperties": {
                "properties": {
                    "sheetId": sheetId,
                    "title": tab,
                    "hidden": True if yes else False,
                },
                "fields": "hidden",
            }
        }
    ]
    if self.flush():
      gexecute(self.service.spreadsheets().batchUpdate(spreadsheetId=self.book_id, body={"requests": req}))
    else:
      self.requests.extend(req)

  def clip(self, tab, shape=(1000, 26)):
    """trim the numer of rows/cols of a page layout"""
    sheetId = self.pages(tab)
    if sheetId is None:
      return
    size = (0, 0)
    meta = gexecute(self.service.spreadsheets().get(spreadsheetId=self.book_id))
    for i, s in enumerate(meta["sheets"]):
      if sheetId == s["properties"]["sheetId"]:
        grid = meta["sheets"][i]["properties"]["gridProperties"]
        size = grid["rowCount"], grid["columnCount"]
        break
    requests = []
    if size[0] > shape[0]:
      del_rows = {
          "range": {
              "sheetId": sheetId,
              "dimension": "ROWS",
              "startIndex": shape[0],
          }
      }
      requests.append({"deleteDimension": del_rows})
    if size[1] > shape[1]:
      del_cols = {
          "range": {
              "sheetId": sheetId,
              "dimension": "COLUMNS",
              "startIndex": shape[1],
          }
      }
      requests.append({"deleteDimension": del_cols})
    if requests:  # and gquota('w'):
      if self.flush():
        gexecute(self.service.spreadsheets().batchUpdate(spreadsheetId=self.book_id, body={"requests": requests}))
      else:
        self.requests.extend(requests)

  def align(self, tab, xy=[]):
    sheetId = self.pages(tab)
    if sheetId is None:
      return
    alignment = {
        "wrapStrategy": "CLIP",
        "horizontalAlignment": "LEFT",
        "verticalAlignment": "MIDDLE",
    }
    rng = {"sheetId": sheetId}
    if xy:
      rng.update({"startRowIndex": xy[0], "endRowIndex": xy[1], "startColumnIndex": xy[2], "endColumnIndex": xy[3]})
    requests = [
        {
            "repeatCell": {
                "range": {"sheetId": sheetId},
                "cell": {"userEnteredFormat": alignment},
                "fields": f"userEnteredFormat({','.join([x for x in alignment])})",
            }
        }
    ]
    if self.flush():
      gexecute(self.service.spreadsheets().batchUpdate(spreadsheetId=self.book_id, body={"requests": requests}))
    else:
      self.requests.extend(requests)

  def border(self, tab, arry):
    sheetId = self.pages(tab)
    if sheetId is None:
      return
    body = json.dumps(
        {
            "updateBorders": {
                "range": {
                    "sheetId": sheetId,
                    "startRowIndex": "ARRY0",
                    "endRowIndex": "ARRY1",
                    "startColumnIndex": "ARRY2",
                    "endColumnIndex": "ARRY3",
                },
                "top": {"style": "SOLID", "width": 1},
                "bottom": {"style": "SOLID", "width": 1},
                "innerHorizontal": {"style": "SOLID", "width": 1},
                "left": {"style": "SOLID", "width": 1},
                "right": {"style": "SOLID", "width": 1},
                "innerVertical": {"style": "SOLID", "width": 1},
            }
        }
    )
    requests = []
    for arr in arry:
      req = body
      for i, number in enumerate(arr):
        req = req.replace(f'"ARRY{i}"', f"{number}")
      requests.append(json.loads(req))
    if self.flush():
      gexecute(self.service.spreadsheets().batchUpdate(spreadsheetId=self.book_id, body={"requests": requests}))
    else:
      self.requests.extend(requests)

  def merge(self, tab, arry):
    sheetId = self.pages(tab)
    if sheetId is None:
      return
    body = json.dumps(
        {
            "mergeCells": {
                "range": {
                    "sheetId": sheetId,
                    "startRowIndex": "ARRY0",
                    "endRowIndex": "ARRY1",
                    "startColumnIndex": "ARRY2",
                    "endColumnIndex": "ARRY3",
                },
                "mergeType": "MERGE_ALL",
            }
        }
    )
    requests = []
    for arr in arry:
      req = body
      for i, number in enumerate(arr):
        req = req.replace(f'"ARRY{i}"', f"{number}")
      requests.append(json.loads(req))
    if self.flush():
      gexecute(self.service.spreadsheets().batchUpdate(spreadsheetId=self.book_id, body={"requests": requests}))
    else:
      self.requests.extend(requests)

  def group(self, tab, rows=[], cols=[], names={}):
    """group in rows and/or cols.example,
    rows = [1,10,20,30] creates 2 groups for lines (1,10) and (20 30)
    names = {'mlbNets':[1,100,0,1]} create named range A2:A101
    """

    sheetId = self.pages(tab)
    if sheetId is None:
      return
    req_row = json.dumps(
        {
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheetId,
                    "dimension": "ROWS",
                    "startIndex": "NUM1",
                    "endIndex": "NUM2",
                },
                "properties": {"hiddenByUser": True},
                # "group": { "collapsed": True  },
                "fields": "hiddenByUser",
            }
        }
    )
    requests = []
    if len(rows) > 1:
      rows = rows[: (len(rows) // 2) * 2]
      for n1, n2 in zip(rows[::2], rows[1::2]):
        req = re.sub(r'"NUM1"', str(n1), re.sub(r'"NUM2"', str(n2), req_row, count=0), count=0)
        requests.append(json.loads(req))
    if len(cols) > 1:
      req_col = re.sub(r'"ROWS"', '"COLUMNS"', req_row, count=0)
      cols = cols[: (len(cols) // 2) * 2]
      for n1, n2 in zip(cols[::2], cols[1::2]):
        req = re.sub(
            r'"NUM1"',
            str(n1),
            re.sub(r'"NUM2"', str(n2), req_col, count=0),
            count=0,
        )
        requests.append(json.loads(req))
    if len(names):  # get exisiting named ranges
      res = gexecute(self.service.spreadsheets().get(spreadsheetId=self.book_id))
      named_rs = res.get("namedRanges", [])
      named_id = {k: None for k in names}
      for k, v in names.items():
        for r in named_rs:
          if k == r["name"]:
            named_id[k] = r["namedRangeId"]
            break
        if len(v) > 3:
          if v[1] is None:
            Range = {
                "sheetId": sheetId,
                "startRowIndex": v[0],
                "startColumnIndex": v[2],
                "endColumnIndex": v[3],
            }
          elif v[3] is None:
            Range = {
                "sheetId": sheetId,
                "startRowIndex": v[0],
                "endRowIndex": v[1],
                "startColumnIndex": v[2],
            }
          else:
            Range = {
                "sheetId": sheetId,
                "startRowIndex": v[0],
                "endRowIndex": v[1],
                "startColumnIndex": v[2],
                "endColumnIndex": v[3],
            }
          if named_id[k]:  # redirect existing named range
            req = {
                "updateNamedRange": {
                    "namedRange": {
                        "namedRangeId": named_id[k],
                        "range": Range,
                    },
                    "fields": "range",
                }
            }
          else:  # create new named range
            req = {"addNamedRange": {"namedRange": {"name": k, "range": Range}}}  # Column B (exclusive, so it's 2)
          requests.append(req)
    if len(requests):
      if self.flush():
        gexecute(self.service.spreadsheets().batchUpdate(spreadsheetId=self.book_id, body={"requests": requests}))
      else:
        self.requests.extend(requests)

  def header(self, tab, rows=[0, 1]):
    """bold font to first row as header"""
    # todo: why this always change color or header to black?
    if rows is None or len(rows) != 2*(len(rows)//2):
      return
    sheetId = self.pages(tab)
    if sheetId is None:
      return None
    # read header content
    rng = f"{tab}!{max(1,rows[0]+1)}:{max(1,rows[1])}"
    res = gexecute(self.service.spreadsheets().values().get(spreadsheetId=self.book_id, range=rng))
    values = res.get("values", [])
    requests = []
    for m,n in zip( rows[0::2], rows[1::2]):
      requests +=[
          {
            "repeatCell": {  # bold
                "range": {
                    "sheetId": sheetId,
                    "startRowIndex": m,
                    "endRowIndex": n,
                },
                "cell": {"userEnteredFormat": {"textFormat": {"bold": True}}},
                "fields": "userEnteredFormat(textFormat)",
            }
          },
          {
              "updateSheetProperties": {  # freeze
                  "properties": {
                      "sheetId": sheetId,
                      "gridProperties": {"frozenRowCount": rows[1] - rows[0]},
                  },
                  "fields": "gridProperties.frozenRowCount",
              }
          },
          {
              "addProtectedRange": {
                  "protectedRange": {
                      "range": {
                          "sheetId": sheetId,
                          "startRowIndex": rows[0],
                          "endRowIndex": rows[1],
                      },
                      # "restrictingRule": { "protectionType": "UNPROTECTED",   "protected": "CONTENT_ONLY"},
                      "description": f"{tab.lower()}Columns",
                      "warningOnly": True,
                  }
              }
          },
        ]

    if requests:
      if self.flush():
        gexecute(self.service.spreadsheets().batchUpdate(spreadsheetId=self.book_id, body={"requests": requests}))
      else:
        self.requests.extend(requests)
    return values if values else None

  def zebra(self, tab, row_chunck):
    """backgroud color rows with alternating grayscale"""
    sheetId = self.pages(tab)
    if sheetId is None:
      return
    color = {0: 1, 1: 0.95}
    body = json.dumps(
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheetId,
                    "startRowIndex": "ARRY0",
                    "endRowIndex": "ARRY1",
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": {
                            "red": "COLOR",
                            "green": "COLOR",
                            "blue": "COLOR",
                        }
                    }
                },
                "fields": "userEnteredFormat(backgroundColor)",
            }
        }
    )
    requests = []
    for i, (k, v) in enumerate(row_chunck.items()):
      req = re.sub(f'"ARRY0"', f"{v[0]+1}", body, count=0)
      req = re.sub(f'"ARRY1"', f"{v[-1]+2}", req, count=0)
      req = re.sub(f'"COLOR"', f"{color[i%2]}", req)
      requests.append(json.loads(req))
    if requests:
      if self.flush():
        gexecute(self.service.spreadsheets().batchUpdate(spreadsheetId=self.book_id, body={"requests": requests}))
      else:
        self.requests.extend(requests)

  def style(self, tab, origin):
    x, y = max(1, origin[0]), max(0, origin[1])
    d = self.read(tab)
    cells = [[]] * len(d)
    for i, (k, v) in enumerate(d.items()):
      cells[i] = [k] + v
      for j, e in enumerate(v):
        if j and not bool(e):
          cells[i][j + 1] = cells[i][j]
    rows, cols = len(cells[0]) - x, len(cells) - y
    border = [x, x + rows, y, y + cols]
    v, w = [], []
    for j, e in enumerate(cells):
      # if not j in [0,2]: continue
      e.append("")
      v.extend([(i, j) for i in range(1, len(e) - 1) if (e[i - 1] != e[i]) and (e[i] == e[i + 1])])
      w.extend([(i, j) for i in range(1, len(e) - 1) if (e[i - 1] == e[i]) and (e[i] != e[i + 1])])
    merge = [[a[0] + x - 1, b[0] + x, a[1] + y, b[1] + y + 1] for (a, b) in zip(v, w)]
    return {
        "border": [border],
        "merge": merge,
        "header": [x - 1, x],
        "align": border,
    }

  def dropdn(self, tab, col_rng={}):
    # create dropdowns for cloumons in col_rng
    sheetId = self.pages(tab)
    if sheetId is None or len(col_rng) < 1:
      return
    requests = []
    for col, rng in filter(lambda e: e[1], col_rng.items()):
      requests.append(
          {
              "setDataValidation": {
                  "rule": {
                      "strict": False,
                      "showCustomUi": True,
                      "condition": {
                          "type": "ONE_OF_RANGE",
                          "values": [{"userEnteredValue": f"={rng}"}],
                      },
                  },
                  "range": {
                      "sheetId": sheetId,
                      "startRowIndex": 1,
                      "startColumnIndex": col,
                      "endColumnIndex": col + 1,
                  },
              }
          }
      )
    if len(requests):
      if self.flush():
        gexecute(self.service.spreadsheets().batchUpdate(spreadsheetId=self.book_id, body={"requests": requests}))
      else:
        self.requests.extend(requests)


class gdoc:

  def __init__(self, url):
    self.service = gservice("docs")
    self.book_id = re.search(".+/document/d/(.+)/edit.*", url).group(1)

    self.doc = gexecute(self.service.documents().get(documentId=self.book_id))
    self.tables = {}
    self.string = ""

  def dump(self, sections={}):
    jstr = json.dumps(self.doc, indent=4, sort_keys=True)
    return jstr

  def get_text(self):
    self.string = self.read_structural_elements(self.doc["body"]["content"])
    return self.string.splitlines()

  def get_table(self, which):
    if not self.tables:
      self.string = self.read_structural_tables(self.doc["body"]["content"])

    if not self.tables:
      return None

    cells = {}
    if type(which) in (str, list):
      which = dict((k, None) for k in re.split(r"\s*,\s*", ",".join(which).strip()))
    if not isinstance(which, dict):
      print("which = {table_name: number_of_header_rows}")
      return

    for tab, ln_headers in which.items():
      for name in self.tables:
        if name.startswith(tab):
          # cells.update({tab: self.tab2cell(self.tables[name],None)})
          cells.update({name: self.tab2cell(self.tables[name], ln_headers)})
          break
    return cells

  def tab2cell(self, table, header_rows=None):
    arry = [["" for j in range(table["columns"])] for i in range(table["rows"])]
    i = 0
    for row in table.get("tableRows"):
      cols = row.get("tableCells")
      j = 0
      for col in cols:
        text = self.read_structural_elements(col.get("content"))
        text = re.sub(r"[\n\x0b\r]", "", text, count=0)
        text = re.sub(r"\s*,\s*", ",", text, count=0)
        arry[i][j] = re.sub(r"\s+", " ", text, count=0).strip()
        j += 1
      i += 1

    if header_rows is None:
      for i in range(1, table["rows"]):
        if arry[i][0] and arry[i][0] != arry[0][0]:
          header_rows = i
          break

    padded = [x[:] for x in arry]
    for i in range(1, header_rows):
      for j in range(table["columns"]):
        if not arry[i][j]:
          padded[i][j] = padded[i - 1][j]
    for i in range(1 + header_rows, table["rows"]):
      for j in range(table["columns"]):
        if not arry[i][j]:
          padded[i][j] = padded[i - 1][j]

    keys = ["" for _ in range(table["columns"])]
    for j in range(table["columns"]):
      keys[j] = " ".join([arry[i][j].strip() for i in range(header_rows)])
    cell = dict((k, []) for k in keys)
    for j, k in enumerate(keys):
      cell[k] = [padded[i][j] for i in range(header_rows, table["rows"])]

    return cell

  def read_paragraph_element(self, e):
    s = e.get("textRun")
    return s.get("content") if s else ""

  def read_structural_elements(self, elements):

    text = ""
    for value in elements:
      if "paragraph" in value:
        elements = value.get("paragraph").get("elements")
        for elem in elements:
          text += self.read_paragraph_element(elem)
      elif "table" in value:
        # The text in table cells are in nested Structural Elements and tables may be
        # nested.
        table = value.get("table")
        for row in table.get("tableRows"):
          cells = row.get("tableCells")
          for cell in cells:
            text += self.read_structural_elements(cell.get("content"))
      elif "tableOfContents" in value:
        # The text in the TOC is also in a Structural Element.
        toc = value.get("tableOfContents")
        text += self.read_structural_elements(toc.get("content"))
    return text

  def read_structural_tables(self, elements):

    text = ""
    for value in elements:
      if "paragraph" in value:
        elements = value.get("paragraph").get("elements")
        for elem in elements:
          text += self.read_paragraph_element(elem)
      elif "table" in value:
        # The text in table cells are in nested Structural Elements and tables may be
        # nested.
        m = re.search(r"\n(.*)$", text.rstrip())
        name = m.group(1).strip() if m else "Table Unamed"
        table = value.get("table")
        self.tables.update({name: table})
        for row in table.get("tableRows"):
          cells = row.get("tableCells")
          for cell in cells:
            text += self.read_structural_tables(cell.get("content"))
      elif "tableOfContents" in value:
        # The text in the TOC is also in a Structural Element.
        toc = value.get("tableOfContents")
        text += self.read_structural_tables(toc.get("content"))
    return text


class gdrive:

  def __init__(self, remote_dir=""):
    self.service = gservice("drive")
    self.remote_dir = self.get_remote(remote_dir)

  def get_remote(self, folder_id):
    if not folder_id:
      return ""
    try:
      r = self.service.files().get(fileId=folder_id, fields="id, name, mimeType").execute()
      if r["mimeType"] == "application/vnd.google-apps.folder":
        return folder_id
      else:
        print(f"The ID '{folder_id}' is not a folder.")
        return ""
    except HttpError as error:
      # Handle error (folder doesn't exist or access is denied)
      if error.resp.status == 404:
        print(f"Folder with ID '{folder_id}' not found.")
      elif error.resp.status == 403:
        print(f"Permission denied for folder with ID '{folder_id}'.")
      else:
        print(f"An error occurred: {error}")
      return ""

  def recent(self, file_id, minutes=3):
    # Fetch the file metadata after upload
    r = self.service.files().get(fileId=file_id, fields="id, name, createdTime").execute()
    created_time = r.get("createdTime")
    created_datetime = datetime.strptime(created_time, "%Y-%m-%dT%H:%M:%S.%fZ")

    current_time = datetime.now(pytz.UTC)
    time_diff = current_time - created_datetime
    recent_threshold = timedelta(minutes=minutes)

    return time_diff <= recent_threshold

  def push(self, local, remote, folder_id=None):
    # use intial remote dir if 3rd arg not given
    remote_dir = self.get_remote(folder_id) or self.remote_dir
    if not remote_dir:
      print(f"gdirve:push: error access remote folder {remote_dir}")
      return

    # make sure local fie exists
    local = local.replace(r"\\", "/")
    if not os.path.isfile(local):
      print(f"gdive:push: local file {local} not found")
      return
    # make sure remote and local has same extention
    _, ext = os.path.splitext(local)
    if ext:
      remote = os.path.splitext(remote)[0] + ext
    _, ext = os.path.splitext(remote)
    # mimeType = gdrive.mime_types.get(ext.lower(), "application/octet-stream")
    mime_type, _ = mimetypes.guess_type(remote)

    # upload
    metadata = {"name": remote, "parents": [remote_dir], "mimeType": mime_type}
    media = MediaFileUpload(local, mimetype=mime_type)
    r = self.service.files().create(body=metadata, media_body=media, ffields="id").execute()
    # check if uploaded recently
    if self.recent(r.get("id"), minutes=3):
      print(f"{local} pushed to {remote_dir} as {remote}")
    pass

  def pull(self, local, remote):
    if not self.get_remote(remote):
      print(f"gdrive:push: not vlaid fildId ? {remote}")
      return
    request = self.service.files().get_media(fileId=remote)
    fh = io.FileIO(local, "wb")  # Open the file in write-binary mode
    downloader = MediaIoBaseDownload(fh, request)

    done = False
    while not done:
      status, done = downloader.next_chunk()
      print(f"Download {int(status.progress() * 100)}%.")

    print(f"File downloaded to: {local}")


class gmail:
  service = gservice("gmail")  # class attribute

  @classmethod
  def send(cls, recipient, subject, message):
    mime = MIMEMultipart()
    mime["to"] = recipient
    mime["subject"] = subject
    mime.attach(MIMEText(message, "plain"))
    rawstr = base64.urlsafe_b64encode(mime.as_bytes()).decode()

    msg = cls.service.users().messages().send(userId="me", body={"raw": rawstr}).execute()
    return msg

def email(receiver, subject, message):
  gmail.send(receiver, subject, message)

if __name__ == "__main__":
  test = None #"gsheet"
  if test == "gsheet":
    url = "https://docs.google.com/spreadsheets/d/1wenue2JMycq36rQlwPz1SO1z77Hhf9PKqRlorzPrPoM/edit?resourcekey=0-tymbYhthaIYsYGjMXtjz6A&gid=2142902699#gid=2142902699"
    book1 = gsheet(url)
    book1.remove(options={"hidden": True})
  if test == "gdrive":
    url = "https://drive.google.com/corp/drive/folders/1FKn-_HR-0d8C5bO5ODshe0Z-JF7Lz6Mg?resourcekey=0-BlLoYSvzlEDANG8PH_tbhw"
    d = gdrive(url)
    d.remote_dir
